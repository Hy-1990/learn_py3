# -*- coding: UTF-8 -*-
# =================================================
# @Project -> File   ：utils -> func
# @IDE    ：PyCharm
# @Author ：至尊宝
# @Date   ：2020/4/27 11:39 上午
# @Desc   ：
# ==================================================

import openpyxl

from excel.bean import Excel_Sheet


def read_excel(path):
    wb = openpyxl.load_workbook(path)

    wb_names = wb.sheetnames

    excel_sheets = []

    # for sheet_name in wb_names:
    #     sheet = wb_names[sheet_name]
    #     x = Excel_Sheet()
    #
    #     max_row = sheet.max_row
    #     max_column = sheet.max_column
    #
    #     x.set_sheet_names(sheet_name)
    #     cells = {}
    #     for row_number in range(max_row):
    #         values = [sheet.cell(row_number, y).value for y in range(max_column)]
    #         cells[row_number] = values


    sheet = wb["nodes"]
    # cell1 = sheet['A1']
    max_column = sheet.max_column
    max_row = sheet.max_row
    print(max_column, max_row)

    for row in range(max_row):
        print("trump:" + str(row))
        for cell in sheet.row:
            print(cell.value, end='\t')
        print('\n')


if __name__ == '__main__':
    read_excel("/Users/huyi/Desktop/er.xlsx")
